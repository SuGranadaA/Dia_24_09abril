#@Librería NumPy

#Importamos la librería para leer .xlsx
import openpyxl

#Importamos la librería de NumPy
import numpy as nmp

#Abrimos el archivo de excel
wob1 = openpyxl.load_workbook('numpy1.xlsx')

#Accedemos a las hojas de cálculo
num1 = wob1['Hoja1']

#Accedemos a las celdas y las transformamos en valores
A1 = num1['A1'].value
B1 = num1['B1'].value
C1 = num1['C1'].value
A3 = num1['A3'].value
B3 = num1['B3'].value
C3 = num1['C3'].value

#Imprimimos los datos
print(A1, B1, C1, A3, B3, C3, "\n")

#Creamos el archivo para almacenar los resultados
res1 = openpyxl.Workbook()

#Asignamos una hoja de cálculo
hoja = res1.active

#Guardamos el archivo
res1.save('res1.xlsx')

#Realizamos una operación, multiplicando C3 por pi
numpi = C3 * nmp.pi
print("Multiplicamos C2 por el número pi: \n", numpi, "\n")

#Realizamos una operación para encontrar las raices de un polinomio formado por A1,C3 y A3
numraiz = nmp.roots([A1, C3, A3])
print("Encontramos las raíces del polinomio: \n", numraiz, "\n")

#Realizamos una operación para establecer un rango en el cual se aumente un número tantas veces
ran = nmp.arange(A1, C1, A3)
print("El rango: \n", ran, "\n")

#Realizamos una operación para crear un número complejo
numcom = complex(A3, B1)
print("Número complejo: \n", numcom, "\n")

#Anexamos la información de las operaciones a la hoja que creamos en celdas específicas
hoja['E1'] = ("Multiplicamos C2 por el número pi: ")
hoja['E2'] = numpi
hoja['F1'] = ("Encontramos las raíces del polinomio: ")
hoja['F2'] = str(numraiz)
hoja['G1'] = ("Obtener el rango: ")
hoja['G2'] = str(ran)
hoja['H1'] = ("Crear un número complejo: ")
hoja['H2'] = str(numcom)

#Guardamos el archivo con las respuestas a los ejercicios
res1.save('res1.xlsx')

#Abrimos el archivo de excel
wob2 = openpyxl.load_workbook('numpy2.xlsx')

#Accedemos a las hojas de cálculo
hojita = wob2['Hoja1']

#Accedemos a las celdas y las transformamos en valores
A22 = hojita['A2'].value
A33 = hojita['A3'].value
A44 = hojita['A4'].value
B22 = hojita['B2'].value
B33 = hojita['B3'].value
B44 = hojita['B4'].value

#Creamos el archivo para almacenar los resultados
sol2 = openpyxl.Workbook()

#Asignamos una hoja de cálculo
hojai = sol2.active

#Guardamos el archivo
sol2.save('res2.xlsx')

#Realizamos una operación, hallando el coseno
cose = nmp.cos(A22)
print("El coseno: \n", cose, "\n")

#Realizamos una operación, hallando el seno de un conjunto
conjunto = [A44, B22, B44]
sen = nmp.sin(conjunto)
print("El seno: \n", sen, "\n")

#Realizamos una operación, sumando un coseno, un seno y un escalar
suma = (nmp.cos(B33) + nmp.sin(A44) + 7)
print("La suma: \n", suma, "\n")

#Realizamos una operación donde convertimos los datos en una matriz
matriz = nmp.matrix([[A22, B22], [B33, A22]])
print("La matriz: \n", matriz, "\n")

#Anexamos la información de las operaciones a la hoja que creamos en celdas específicas
hojai['E1'] = ("Calcular coseno: ")
hojai['E2'] = cose
hojai['F1'] = ("Calcular seno: ")
hojai['F2'] = str(sen)
hojai['G1'] = ("Suma de seno y coseno + 7: ")
hojai['G2'] = (suma)
hojai['H1'] = ("Creación de matriz: ")
hojai['H2'] = str(matriz)

#Guardamos el archivo con las respuestas a los ejercicios
sol2.save('res2.xlsx')
